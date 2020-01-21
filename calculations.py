from main import main
from estimates import *
import numpy as np
import openpyxl
from distributions import *


def N():
    result_table = []
    for i in range(200):
        dictionary, que, dev = main(2000, 0)
        result_table.append([dictionary['p'], dictionary['Tq'], dictionary['Ts'], dictionary['Nq'], dictionary['Ns'],
                             dictionary['Ca'], dictionary['Cr']])
    m = []
    d = []
    for i in range(7):
        temp = []
        for j in range(len(result_table)):
            temp.append(result_table[j][i])
        t1, t2 = estimates(temp)
        m.append(t1)
        d.append(t2)
    result_table.append(m)
    result_table.append(d)
    eps = [i * 0.05 for i in m]
    result_table.append(eps)
    u = 1.96
    result_table.append([u ** 2 * d[i] / eps[i] ** 2 for i in range(7)])

    wb = openpyxl.Workbook()
    wb.create_sheet(title='Первый лист', index=0)
    sheet = wb['Первый лист']
    sheet.append(result_table[len(result_table) - 4])
    sheet.append(result_table[len(result_table) - 3])
    sheet.append(result_table[len(result_table) - 2])
    sheet.append(result_table[len(result_table) - 1])
    wb.save('n.xlsx')


m_a = 10
x1_min = 3
x1_max = 5
x2_min = 20
x2_max = 30
x3_min = 20
x3_max = 34
factor_plan = [[5, 25, 34],
                   [x1_min, x2_min, x3_min],
                   [x1_max, x2_min, x3_min],
                   [x1_min, x2_max, x3_min],
                   [x1_max, x2_max, x3_min],
                   [x1_min, x2_min, x3_max],
                   [x1_max, x2_min, x3_max],
                   [x1_min, x2_max, x3_max],
                   [x1_max, x2_max, x3_max],
                   [3, 22, 25],
                   [3, 22, 30],
                   [3, 24, 25],
                   [3, 24, 30],
                   [3, 26, 25],
                   [3, 26, 30],
                   [4, 22, 25],
                   [4, 22, 30],
                   [4, 24, 25],
                   [4, 24, 30],
                   [4, 26, 25],
                   [4, 26, 30]]
# factor_plan = [[x1_min, x2_min, x3_min],
#                    [x1_max, x2_min, x3_min],
#                    [x1_min, x2_max, x3_min],
#                    [x1_max, x2_max, x3_min],
#                    [x1_min, x2_min, x3_max],
#                    [x1_max, x2_min, x3_max],
#                    [x1_min, x2_max, x3_max],
#                    [x1_max, x2_max, x3_max]]

# factor_plan = [[5, 22, 25],
#                [5, 24, 25],
#                [5, 24, 30],
#                [5, 26, 25],
#                [5, 26, 30],
#                [5, 28, 25]]


def Res():
    n = 135

    R = []
    wb = openpyxl.Workbook()
    for i in range(len(factor_plan)):
        temp_table = []
        m = []
        wb.create_sheet(title=str(i), index=0)
        sheet = wb[str(i)]

        for j in range(n):
            dictionary, que, dev = main(2000, 0, factor_plan[i][0], factor_plan[i][2],
                                                       factor_plan[i][1], m_a)
            temp_table.append([dictionary['p'], dictionary['Tq'], dictionary['Ts'],
                               dictionary['Nq'], dictionary['Ns'], dictionary['Ca'], dictionary['Cr']])
            sheet.append([dictionary['p'], dictionary['Tq'], dictionary['Ts'],
                          dictionary['Nq'], dictionary['Ns'], dictionary['Ca'], dictionary['Cr']])
        for t in range(7):
            temp = []
            for l in range(len(temp_table)):
                temp.append(temp_table[l][t])
            m.append(round(sum(temp) / len(temp), 3))
        R.append(m)
        sheet.append(m)
    # wb.save('R1.xlsx')

    wb = openpyxl.Workbook()
    wb.create_sheet(title='Первый лист', index=0)
    sheet = wb['Первый лист']
    for row in range(len(R)):
        for col in range(len(R[row])):
            cell = sheet.cell(row=row + 1, column=col + 1)
            cell.value = str(R[row][col])

    wb.save('R.xlsx')

    return R


def E(R):
    e1 = []
    e2 = []
    e3 = []
    e12 = []
    e13 = []
    e23 = []
    e123 = []
    for k in range(7):
        e1.append((R[1][k] - R[0][k] + R[3][k] - R[2][k] + R[5][k] -
                   R[4][k] + R[7][k] - R[6][k]) / 4)
        e2.append((R[2][k] - R[0][k] + R[3][k] - R[1][k] + R[6][k] -
                   R[4][k] + R[7][k] - R[5][k]) / 4)
        e3.append((R[4][k] - R[0][k] + R[5][k] - R[1][k] + R[6][k] -
                   R[2][k] + R[7][k] - R[3][k]) / 4)

        e12.append(1 / 2 * ((R[3][k] - R[2][k] + R[7][k] - R[6][k]) / 2 - (R[1][k] - R[0][k] + R[5][k] - R[4][k]) / 2))
        e13.append(1 / 2 * ((R[5][k] - R[4][k] + R[7][k] - R[6][k]) / 2 - (R[1][k] - R[0][k] + R[3][k] - R[2][k]) / 2))
        e23.append(1 / 2 * ((R[6][k] - R[4][k] + R[7][k] - R[5][k]) / 2 - (R[2][k] - R[0][k] + R[3][k] - R[1][k]) / 2))

        e123.append(1 / 2 * ((R[7][k] - R[6][k] - R[5][k] + R[4][k]) / 2 - (R[3][k] - R[2][k] - R[1][k] + R[0][k]) / 2))

    wb = openpyxl.Workbook()
    wb.create_sheet(title='Первый лист', index=0)
    sheet = wb['Первый лист']
    sheet.append(e1)
    sheet.append(e2)
    sheet.append(e3)
    sheet.append(e12)
    sheet.append(e13)
    sheet.append(e23)
    sheet.append(e123)

    wb.save('e.xlsx')


def table_A(R):
    a = []
    A = np.array(
        [[1, x1_min, x2_min, x3_min, x1_min * x2_min, x1_min * x3_min, x2_min * x3_min, x1_min * x2_min * x3_min],
         [1, x1_max, x2_min, x3_min, x1_max * x2_min, x1_max * x3_min, x2_min * x3_min, x1_max * x2_min * x3_min],
         [1, x1_min, x2_max, x3_min, x1_min * x2_max, x1_min * x3_min, x2_max * x3_min, x1_min * x2_max * x3_min],
         [1, x1_max, x2_max, x3_min, x1_max * x2_max, x1_max * x3_min, x2_max * x3_min, x1_max * x2_max * x3_min],
         [1, x1_min, x2_min, x3_max, x1_min * x2_min, x1_min * x3_max, x2_min * x3_max, x1_min * x2_min * x3_max],
         [1, x1_max, x2_min, x3_max, x1_max * x2_min, x1_max * x3_max, x2_min * x3_max, x1_max * x2_min * x3_max],
         [1, x1_min, x2_max, x3_max, x1_min * x2_max, x1_min * x3_max, x2_max * x3_max, x1_min * x2_max * x3_max],
         [1, x1_max, x2_max, x3_max, x1_max * x2_max, x1_max * x3_max, x2_max * x3_max, x1_max * x2_max * x3_max]])

    for i in range(7):
        b = np.array([[R[0][i]], [R[1][0]], [R[2][i]], [R[3][i]], [R[4][i]], [R[5][i]], [R[6][i]], [R[7][i]]])
        t = np.linalg.inv(A).dot(b)
        print(t)
        a.append(t)
    wb = openpyxl.Workbook()
    wb.create_sheet(title='Первый лист', index=0)
    sheet = wb['Первый лист']
    for i in range(len(a)):
        for j in range(len(a[i])):
            cell = sheet.cell(row=j + 1, column=i + 1)
            cell.value = str(round(a[i][j][0], 3))

    y = []
    for i in a:
        y.append(i[0] + i[1] * 5 + i[2] * 25 + i[3] * 34 + i[4] * 5 * 25 + i[5] * 5 * 34 + i[6] * 25 * 34 + i[
            7] * 5 * 34 * 25)

    for row in range(len(y)):
        for col in range(len(y[row])):
            cell = sheet.cell(row=row + 11, column=col + 1)
            cell.value = str(y[row][col])

    wb.save('a.xlsx')


def price(R):
    c1 = 500000
    c2 = 50000
    c3 = 300
    c4 = 0.08
    c5 = 0.093

    En = 0.15
    T = 25 * 10 ^ 7
    I = []

    for i in range(len(R)):
        s = factor_plan[i][0]
        I.append(
            En * c1 * s + c2 * (R[i][4] - R[i][3]) + c3 * (s - R[i][4] + R[i][3]) + c4 * T * (
                    1 / m_a - R[i][5]) + c5 * T * R[i][3])
    wb = openpyxl.Workbook()
    wb.create_sheet(title='Первый лист', index=0)
    sheet = wb['Первый лист']
    for row in range(len(I)):
        cell = sheet.cell(row=row + 1, column=1)
        cell.value = str(round(I[row], 2))

    wb.save('price.xlsx')
    # for t in I:
    #     print(t)


# distr1 = []
# for i in range(1000):
#     distr1.append(uni_dist_1())
# m, d = estimates(distr1)
# scatter_plot(distr1)
# correlation(distr1, m, d)
# bar_graph(distr1, m)
N()
# table = Res()
# E(table)
# table_A(table)
# table = Res()
# price(table)
